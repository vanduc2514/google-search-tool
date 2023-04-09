import logging
from copy import copy
from datetime import datetime

import openpyxl

from utils import now_as


class ExcelReporter:

    def __init__(self, workbook_path, sheet_name):
        self._workbook_path = workbook_path
        self._sheet_name = sheet_name
        self._report_date = now_as("%d/%m/%Y")

    def export_results(self, results):
        workbook, sheet, last_row = self.__open__()
        start_row = last_row + 1
        for search_result in results:
            report_row = __create_report_row__(sheet,
                                               row=start_row + 1,
                                               last_row=last_row)
            __fill_data__(report_row, self._report_date, search_result)
        # Save the changes
        workbook.save(self._workbook_path)
        logging.info(f"Exported Report to {self._workbook_path} using Sheet {self._sheet_name}")

    def __open__(self):
        workbook = openpyxl.load_workbook(self._workbook_path)
        sheet = workbook[self._sheet_name]
        # Find the last row which have content
        last_row = sheet.max_row
        while sheet.cell(row=last_row, column=1).value is None and last_row > 1:
            last_row -= 1
        logging.debug(f"Last Row has content: {last_row}")
        return workbook, sheet, last_row


def __create_report_row__(sheet, row, last_row):
    sheet.row_dimensions[row].height = 30
    report_row = []
    for last_cell in sheet[last_row]:
        new_cell = sheet.cell(row=row, column=last_cell.column)
        __copy_style__(new_cell, last_cell)
        report_row.append(last_cell)
    return report_row


def __copy_style__(cell, other_cell):
    cell.alignment = copy(other_cell.alignment)
    cell.font = copy(other_cell.font)
    cell.border = copy(other_cell.border)
    cell.fill = copy(other_cell.fill)


def __fill_data__(row, report_date, search_result):
    row[0].value = report_date
    row[1].value = search_result.keyword
    url_cell_ = row[2]
    if search_result.result is None:
        url_cell_.value = f"No Result Found in: {search_result.site}"
    else:
        url_cell_.value = search_result.result.link
        url_cell_.hyperlink = search_result.result.link
